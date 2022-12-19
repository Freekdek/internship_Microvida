# Created by Freek de Kreek, 14-9-2022
# For each file:
# o Convert file name to MWGS number only and put it in dictionary
# o MWGS number found 
# o Save GLIMS ID by adding to dictionary value (list)
# o Store isolate number by adding to dictionary value (list)
# o Write file name  GLIMS ID + “1” + isolate number + “wgs.pdf”

# Latest update: 16-9-2022, added file already exist check before copying and renaming the sample reports and changed file_col to 1

import os
import argparse
import openpyxl
import shutil
import timeit

start = timeit.default_timer() # starts timer

# initialize parser
parser = argparse.ArgumentParser()
# adding optional argument
parser.add_argument("-i", "--Input", help="Show Input sample reports directory")
parser.add_argument("-d", "--Database", help="Show Input WGS overzicht BRMO surveillance")
parser.add_argument("-o", "--Output", help="Show Output directory")
parser.add_argument("-v", "--Verbose", help="Show more information")
# read arguments from command line
args = parser.parse_args()
input = args.Input
database = args.Database
output = args.Output
verbose = args.Verbose

if verbose == "True":
    debug = True
else:
    debug = False

if input and database and output:
    file_list = os.listdir(input) # Creates a list of all files in directory (specified at input)
    # define variable to load the wookbook
    workbook = openpyxl.load_workbook(database)
    # define variable to read the active sheet:
    sheet = workbook.worksheets[0]

    for file in file_list: # every sample report is iterated
        if file.endswith(".pdf"):
            print(f"{file} is being renamed...")
            file_path = os.path.join(input + file) # path of the file (one of the sample reports)
            if debug:
                print(file + " is being renamed...")
            split1 = file.split("Report_")
            split2 = split1[-1]
            mwgs = split2.replace(".pdf", "") # file name changed from Report_MWGS_ID.pdf to MWGS_ID
            if debug:
                print(mwgs)
            # row and collumn file to be read
            file_row = 1
            file_col = 1 # changed from 2 to 1, due to the new Nextera template (BRMO overzicht)
            for row in range(1, sheet.max_row + 1):  # every row is searched for the MWGS IDs
                if sheet.cell(file_row, file_col).value == mwgs:
                    GLIMS = sheet.cell(file_row, file_col + 2).value
                    isolate = sheet.cell(file_row, file_col + 3).value
                    new_name = GLIMS + "1" + str(isolate) + "wgs.pdf" # creates the correct new name
                    destination = output + new_name # adds the new name to the output folder
                    if debug:
                        print(f"found mwgs id: {mwgs} with GLIMS: {GLIMS} and isolate: {isolate}")
                        print(destination)
                    if os.path.exists(destination):  # checks whether the file already exists (sample report in Glims ID format)
                        print(f"{new_name} already exists, aborting copying...")
                    else:
                        shutil.copyfile(file_path, destination)  # copies the sample report and places it in the output folder with the new and correct name, added 16-9-2022
                file_row += 1 # goes to the next row
            file_row = 0 # resets starting row

stop = timeit.default_timer() # ends timer

print('Time: ', stop - start) # prints script run duration