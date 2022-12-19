# Created by Freek de Kreek, 26-8-2022 #

# Reads all Excel files (qc trim reports) of specified  directory (Input) and collects all the required values (in Output):
# Any nucleotide, Count, N50, Maximum, b Reads, Contigs, Matched
# Every row is read, checks for the afore mentioned values and when found writes the values in the corresponding collumn
# Duplicates are not allowed

# Latest update: 30-8-2022, corrected index for writing the values and testing only read the third sheet.
# Latest update: 6-9-2022, removed read_only=True in openpyxl.load_workbook to  line 69
# Latest update: 8-9-2022, added wgs and mwgs_id to change the name from MWGS_id_TrimReport.xlsx to just the MWGS_id


import os
import openpyxl
import argparse
import timeit
import shutil

start = timeit.default_timer() # starts timer


# initialize parser
parser = argparse.ArgumentParser()
# adding optional argument
parser.add_argument("-o", "--Output", help="Show Output (QC assembly template)")
parser.add_argument("-i", "--Input", help="Show Input of QC trim reports directory")
# read arguments from command line
args = parser.parse_args()
input = args.Input
target = args.Output

if input and target:
    # os.system("chmod 777 " + target) # grants permission to all files within the specified folder (input) to be read and written
    # os.system("chmod 755 " + input)
    # os.system("chmod 755 " + input + "MWGS*.xlsx")  # grants permission to all files within the specified folder (input) to be read
    file_list = os.listdir(input) # creates a list of all files in directory (specified at input)
    # creates a copy of the qc assembly template specified at output
    if target.endswith('1.xlsx'):
        copy = target.replace("1.xlsx", ".xlsx")
        shutil.copyfile(target, copy)
    elif target.endswith('_.xlsx'):
        copy = target.replace(".xlsx", "1.xlsx")
        shutil.copyfile(target, copy)
    else:
        print("check if QC assembly template has the correct name: it should end with ).xlsx or 1.xlsx")
        copy = target.replace(".xlsx", "naming_error.xlsx")
        shutil.copyfile(target, copy)
    # reads the QC template (which is the specified output)
    workbook = openpyxl.load_workbook(target)
    worksheet = workbook.worksheets[0]
    # start from the first cell, rows are and columns are not zero indexed (output)
    row_result = 2
    col_result = 0
    # row and collumn file to be read (input)
    file_row = 2
    file_col = 1

    for file in file_list:
        if file.endswith('.xlsx'):
            file_path = os.path.join(input + file) # path of the file (one of the qc reports)
            print(file + " is loading")
            # write MWGS_id
            mwgs = file.split("_TrimReport")
            mwgs_id = mwgs[0]
            worksheet.cell(row_result, col_result + 1, mwgs_id) # writes down the MWGS id under the first collumn
            # define variable to load the wookbook
            wookbook = openpyxl.load_workbook(file_path)
            # define variable to read the active sheet
            try:
                sheet = wookbook.worksheets[2]  # added 30-8-2022, should only read sheet 3 (saving time)

                dict = {
                    "Any nucleotide (N)": True,
                    "Count": True,
                    "N50": True,
                    "Maximum": True,
                    "Total": True,
                    "Reads": True,
                    "Contigs": True,
                    "Matched": True
                }  # creates a dictionary for all values and if they are duplicate (=False) or not (=True)

                for row in range(1, sheet.max_row+1): # every row is searched for the required values
                    cell_value = sheet.cell(row=file_row, column=file_col).value # current cell that is read
                    if cell_value in dict.keys() and dict[cell_value] == True: # this line checks if the value of the read cell is within the dictionary, so that when true it can be further analyzed
                        if sheet.cell(row=file_row, column=file_col).value == "Any nucleotide (N)" and dict["Any nucleotide (N)"] == True:
                            any_N = sheet.cell(row=file_row, column=file_col + 1).value # value of Any nucleotide is saved
                            worksheet.cell(row_result, col_result + 6).value = any_N # write the any_N value in the second collumn
                            dict["Any nucleotide (N)"] = False
                        elif sheet.cell(row=file_row, column=file_col).value == "Count" and dict["Count"] == True:
                            Count = sheet.cell(row=file_row, column=file_col + 1).value # value of number of contigs is saved
                            worksheet.cell(row_result, col_result + 7).value = Count
                            dict["Count"] = False
                        elif sheet.cell(row=file_row, column=file_col).value == "N50" and dict["N50"] == True:
                            N50 = sheet.cell(row=file_row, column=file_col + 1).value  # value of N50 is saved
                            worksheet.cell(row_result, col_result + 8).value = N50
                            dict["N50"] = False
                        elif sheet.cell(row=file_row, column=file_col).value == "Maximum" and dict["Maximum"] == True:
                            Maximum = sheet.cell(row=file_row, column=file_col + 1).value # value of max contig length is saved
                            worksheet.cell(row_result, col_result + 9).value = Maximum
                            dict["Maximum"] = False
                        elif sheet.cell(row=file_row, column=file_col).value == "Total" and dict["Total"] == True:
                            Total = sheet.cell(row=file_row, column=file_col + 1).value # value of total basepairs is saved
                            worksheet.cell(row_result, col_result + 10).value = Total
                            dict["Total"] = False
                        elif sheet.cell(row=file_row, column=file_col).value == "Reads" and dict["Reads"] == True:
                            Reads = sheet.cell(row=file_row, column=file_col + 1).value # value of read counts is saved
                            Reads_avg = sheet.cell(row=file_row, column=file_col + 2).value  # value of average read length is saved
                            worksheet.cell(row_result, col_result + 11).value = Reads
                            worksheet.cell(row_result, col_result + 12).value = Reads_avg
                            dict["Reads"] = False
                        elif sheet.cell(row=file_row, column=file_col).value == "Contigs" and dict["Contigs"] == True:
                            Contigs = sheet.cell(row=file_row, column=file_col + 3).value # value of contig total bases is saved
                            worksheet.cell(row_result, col_result + 13).value = Contigs
                            dict["Contigs"] = False
                        elif sheet.cell(row=file_row, column=file_col).value == "Matched" and dict["Matched"] == True:
                            Matched = sheet.cell(row=file_row, column=file_col + 1).value # value of count matched is saved
                            worksheet.cell(row_result, col_result + 14).value = Matched
                            dict["Matched"] = False

                    file_row += 1 # next row to read
            except IndexError:
                print(f"{mwgs_id} failed, Trimreport does not contain the results sheet (sheet 3 in the excel file) \n"
                      f"no QC results can be filled in for {mwgs_id} :(")
            row_result += 1
            file_row = 1
            wookbook.close()
    workbook.save(target)
    workbook.close()

stop = timeit.default_timer() # ends timer

print('Time: ', stop - start) # prints script run duration